import openpyxl as pyxl
import ctypes
from sys				import(argv, exit)
from PyQt5.Qt			import(QApplication,
							   QMainWindow)
from PyQt5.QtCore		import(Qt)
from PyQt5.QtWidgets	import(QGridLayout,
							   QHBoxLayout,
							   QVBoxLayout,
							   QFrame,
							   QGroupBox,
							   QLabel,
							   QPushButton,
							   QLineEdit,
							   QTableWidget,
							   QTableWidgetItem,
							   QTabWidget,
							   QFileDialog)




class PyExcel(QMainWindow):
	
	if " O M A R " :
		print(" ...                   ")
		print(" Hello to PyExcel ! .. ")
		print(" # O M A R | 2 0 0 3 # ")
		print(" ...                   ")
		print()


	def __init__(self, Developer):
		
		QMainWindow.__init__(self)

		self.Developer = Developer

		self.screen_width  = ctypes.windll.user32.GetSystemMetrics(0)
		self.screen_height = ctypes.windll.user32.GetSystemMetrics(1)
		self.width  = int(self.screen_width *0.70)
		self.height = int(self.screen_height*0.75)

		self.setWindowTitle('PyExcel')
		self.setGeometry(int((self.screen_width-self.width)/2),
						 int((self.screen_height-self.height)/2),
						 self.width,
						 self.height)
		self.setFixedSize(self.width,self.height)
		
		self.Open_Excel_File()
		self.GUI()
		self.Connect()
		self.Download_Table()
		self.Current()




	def GUI(self):

		self.Frame = QFrame(self)
		self.Frame.setGeometry(10,10,self.width-20,self.height-20)
		self.Frame.setLineWidth(1)
		self.Frame.setFrameShape(QFrame.StyledPanel | QFrame.Raised)
		self.Grid = QGridLayout(self.Frame)

		self.TabWidget = QTabWidget(self.Frame)
		self.TabWidget.setFixedSize(self.width-510,self.height-60)
		for index , name in enumerate(self.SheetNames):
			print(f"Sheet {index} : {name}.")
			self.TabWidget.addTab(QTableWidget(self.RowsCount[index],self.ColumnsCount[index]),name)
		self.Grid.addWidget(self.TabWidget,0,0, Qt.AlignLeft | Qt.AlignCenter)


		self.Frame2 = QFrame(self)
		self.Frame2.setFixedSize(450,self.height-40)
		self.Grid2 = QGridLayout(self.Frame2)
		self.Grid2.setContentsMargins(10,100,10,100)
		self.Grid2.setVerticalSpacing(40)
		self.Grid.addWidget(self.Frame2,0,1)

		self.GBall_population = QGroupBox("Find All Population")
		self.Ball_population = QPushButton("Find")
		self.Lall_population = QLabel("All Population!")
		self.GBall_population_Grid = QHBoxLayout(self.GBall_population)
		self.GBall_population_Grid.addWidget(self.Ball_population,1)
		self.GBall_population_Grid.addWidget(self.Lall_population,2)
		self.Grid2.addWidget(self.GBall_population,0,0)

		self.GBmax_and_min = QGroupBox("Find Maximum And Minimum Population")
		self.Lmax = QLabel("Maximum Population : ")
		self.Lmin = QLabel("Minimum Population : ")
		self.Bmax_and_min = QPushButton("Find")
		self.GBmax_and_min_Grid = QVBoxLayout(self.GBmax_and_min)
		self.GBmax_and_min_Grid.addWidget(self.Lmax)
		self.GBmax_and_min_Grid.addWidget(self.Lmin)
		self.GBmax_and_min_Grid.addWidget(self.Bmax_and_min)
		self.Grid2.addWidget(self.GBmax_and_min,1,0)

... (175 lines left)
Collapse
Excel_APP.py
9 KB
﻿
import openpyxl as pyxl
import ctypes
from sys				import(argv, exit)
from PyQt5.Qt			import(QApplication,
							   QMainWindow)
from PyQt5.QtCore		import(Qt)
from PyQt5.QtWidgets	import(QGridLayout,
							   QHBoxLayout,
							   QVBoxLayout,
							   QFrame,
							   QGroupBox,
							   QLabel,
							   QPushButton,
							   QLineEdit,
							   QTableWidget,
							   QTableWidgetItem,
							   QTabWidget,
							   QFileDialog)




class PyExcel(QMainWindow):
	
	if " O M A R " :
		print(" ...                   ")
		print(" Hello to PyExcel ! .. ")
		print(" # O M A R | 2 0 0 3 # ")
		print(" ...                   ")
		print()


	def __init__(self, Developer):
		
		QMainWindow.__init__(self)

		self.Developer = Developer

		self.screen_width  = ctypes.windll.user32.GetSystemMetrics(0)
		self.screen_height = ctypes.windll.user32.GetSystemMetrics(1)
		self.width  = int(self.screen_width *0.70)
		self.height = int(self.screen_height*0.75)

		self.setWindowTitle('PyExcel')
		self.setGeometry(int((self.screen_width-self.width)/2),
						 int((self.screen_height-self.height)/2),
						 self.width,
						 self.height)
		self.setFixedSize(self.width,self.height)
		
		self.Open_Excel_File()
		self.GUI()
		self.Connect()
		self.Download_Table()
		self.Current()




	def GUI(self):

		self.Frame = QFrame(self)
		self.Frame.setGeometry(10,10,self.width-20,self.height-20)
		self.Frame.setLineWidth(1)
		self.Frame.setFrameShape(QFrame.StyledPanel | QFrame.Raised)
		self.Grid = QGridLayout(self.Frame)

		self.TabWidget = QTabWidget(self.Frame)
		self.TabWidget.setFixedSize(self.width-510,self.height-60)
		for index , name in enumerate(self.SheetNames):
			print(f"Sheet {index} : {name}.")
			self.TabWidget.addTab(QTableWidget(self.RowsCount[index],self.ColumnsCount[index]),name)
		self.Grid.addWidget(self.TabWidget,0,0, Qt.AlignLeft | Qt.AlignCenter)


		self.Frame2 = QFrame(self)
		self.Frame2.setFixedSize(450,self.height-40)
		self.Grid2 = QGridLayout(self.Frame2)
		self.Grid2.setContentsMargins(10,100,10,100)
		self.Grid2.setVerticalSpacing(40)
		self.Grid.addWidget(self.Frame2,0,1)

		self.GBall_population = QGroupBox("Find All Population")
		self.Ball_population = QPushButton("Find")
		self.Lall_population = QLabel("All Population!")
		self.GBall_population_Grid = QHBoxLayout(self.GBall_population)
		self.GBall_population_Grid.addWidget(self.Ball_population,1)
		self.GBall_population_Grid.addWidget(self.Lall_population,2)
		self.Grid2.addWidget(self.GBall_population,0,0)

		self.GBmax_and_min = QGroupBox("Find Maximum And Minimum Population")
		self.Lmax = QLabel("Maximum Population : ")
		self.Lmin = QLabel("Minimum Population : ")
		self.Bmax_and_min = QPushButton("Find")
		self.GBmax_and_min_Grid = QVBoxLayout(self.GBmax_and_min)
		self.GBmax_and_min_Grid.addWidget(self.Lmax)
		self.GBmax_and_min_Grid.addWidget(self.Lmin)
		self.GBmax_and_min_Grid.addWidget(self.Bmax_and_min)
		self.Grid2.addWidget(self.GBmax_and_min,1,0)

		self.GBcity_name = QGroupBox("Find Population Of Specific City")
		self.LEcity_name = QLineEdit()
		self.LEcity_name.setPlaceholderText("Enter City Name")
		self.Bcity_name = QPushButton("Find")
		self.Lcity_name = QLabel("City Population")
		self.GBcity_name_Grid = QGridLayout(self.GBcity_name)
		self.GBcity_name_Grid.addWidget(self.LEcity_name,0,0,1,3)
		self.GBcity_name_Grid.addWidget(self.Bcity_name,1,0)
		self.GBcity_name_Grid.addWidget(self.Lcity_name,1,1,1,2)
		self.Grid2.addWidget(self.GBcity_name,2,0)

		for label in [self.Lall_population,self.Lmax,self.Lmin,self.Lcity_name]:
			label.setWordWrap(True)
		
		self.Style_Sheet()


	def Style_Sheet(self):
		self.setStyleSheet("""
						QGroupBox {
							font-size    : 15px;
							border       : 2px ridge gray;
							border-radius: 5px;
							padding-top  : 20px;
						}
						QGroupBox::title {
							background-color   : transparent;
							subcontrol-position: top ;
							padding-top        : 10px;
						}
						QPushButton {
							font-size: 15px;
							max-width: 60px;
						}
						QLabel {
							font-size: 15px;
						}
						QLineEdit {
							font-size: 15px;
						}
						QTableWidget {
							font-size: 18px;
						}
						QHeaderView::section:horizontal {
							font-size : 15px;
							min-width : 300px;
							height    : 25px;
						}
						QHeaderView::section:vertical {
							font-size: 15px;
							width: 35px;
						}
						QTabBar::tab {
							font-size: 15px;
							min-width: 150px;
						}
						""")
		

	def Current(self):
		self.CurrentIndex = self.TabWidget.currentIndex()
		self.CurrentSheet = self.SheetNames[self.CurrentIndex]


	def Connect(self):
		self.TabWidget.currentChanged.connect(lambda: self.Current())
		self.Ball_population.clicked .connect(lambda: self.All_Population     ())
		self.Bmax_and_min   .clicked .connect(lambda: self.Maximum_And_Minimum())
		self.Bcity_name     .clicked .connect(lambda: self.City_Name          ())
	
	
	def Open_Excel_File(self):
		self.FileName      = QFileDialog.getOpenFileName(self,"Open Excel File","","Excel Files (*.xlsx)")[0]
		self.WorkBook      = pyxl.load_workbook(self.FileName,True)
		self.SheetNames    = self.WorkBook.sheetnames
		self.WorkSheets    = self.WorkBook.worksheets
		self.RowsCount     = [sheet.max_row    for sheet in self.WorkSheets]
		self.ColumnsCount  = [sheet.max_column for sheet in self.WorkSheets]
	

	def Download_Table(self):	
		for index , name in enumerate(self.SheetNames):	
			table = self.TabWidget.widget(index)
			sheet = self.WorkBook[name]
			for row in range(sheet.max_row):
				item = QTableWidgetItem(str(sheet[f'A{row+1}'].value))
				item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
				table.setItem(row,0,item)
			for row in range(sheet.max_row):
				item = QTableWidgetItem(str(sheet[f'B{row+1}'].value))
				item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
				table.setItem(row,1,item)
			table.resizeColumnsToContents()
	



	def All_Population(self):
		sheet = self.WorkBook[self.CurrentSheet]	
		all_pop = 0	
		
		for row in range(sheet.max_row):
			all_pop += int(sheet[f'B{row+1}'].value)
		self.Lall_population.setText(f"Total population in {self.CurrentSheet} =  {all_pop}.")



	def Maximum_And_Minimum(self):
		sheet = self.WorkBook[self.CurrentSheet]
		
		for row in range(sheet.max_row):
			if not row:
				max_pop  = sheet[f'B{row+1}'].value
				min_pop  = sheet[f'B{row+1}'].value
				max_city = sheet[f'A{row+1}'].value
				min_city = sheet[f'A{row+1}'].value
				continue
			if  max_pop  < sheet[f'B{row+1}'].value:
				max_pop  = sheet[f'B{row+1}'].value
				max_city = sheet[f'A{row+1}'].value
			if  min_pop  > sheet[f'B{row+1}'].value:
				min_pop  = sheet[f'B{row+1}'].value
				min_city = sheet[f'A{row+1}'].value
		
		self.Lmax.setText(f"Maximum population in {self.CurrentSheet} is {max_city} with population of : {max_pop}.")
		self.Lmin.setText(f"Minimum population in {self.CurrentSheet} is {min_city} with population of : {min_pop}.")



	def City_Name(self):
		city_name = self.LEcity_name.text().strip().lower()
		if not city_name:
			self.Lcity_name.setText(f"Enter a city in {self.CurrentSheet} to find its population.")
			return
	
		sheet = self.WorkBook[self.CurrentSheet]
		found = False

		for row in range(sheet.max_row):
			if sheet[f"A{row+1}"].value.strip().lower() == city_name:
				found = True
				self.Lcity_name.setText(f"Population of {city_name} = {sheet[f'B{row+1}'].value}.")
				self.TabWidget.widget(self.CurrentIndex).selectRow(row)
				break

		if not found: self.Lcity_name.setText(f"({city_name}) is not exist in {self.CurrentSheet}.")





if __name__ == '__main__':

	root = QApplication(argv)
	root.setStyle('Fusion')
	
	App = PyExcel(" O M A R ")
	App.show()

	exit(root.exec_())





if PyExcel.Developer == " O M A R " :

	Omar = "               DEVELOPER               "
	PyXl = "                PROJECT                "
	File = "E:/Programming/Python/Alshoki/pyxl.xlsx"

